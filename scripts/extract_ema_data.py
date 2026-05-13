#!/usr/bin/env python3
"""Extract and merge EMA data from DANE Excel files into a master CSV.

Reads every ``.xlsx`` file in ``data/raw/``, extracts three time series
for the **Caribe** region from each file, deduplicates by keeping the
latest revision for each date, and writes a single CSV to
``data/processed/master_tourism_series.csv``.

Source sheets (names vary slightly between years, discovered by keyword):

    *Occupancy*
        contains ``"Porc Mens Ocupacion.reg"``
        (e.g. ``5.2 Porc Mens Ocupacion.reg`` or
        ``4.2 Porc Mens Ocupacion.reg``)

    *Supply & Demand*
        contains ``"Ind.Mes oferta.demanda"``
        (e.g. ``8.1 Ind.Mes oferta.demanda`` or
        ``7.1 Ind.Mes oferta.demanda``)

    *Real Income*
        contains ``"1.1"`` and ``"Ing.real"``
        (always ``1.1 V.A Ing.real``)

Output columns:

    - ``Date``                  datetime (first day of month)
    - ``Ocupacion_Caribe``      occupancy percentage
    - ``Hab_Disponibles_Caribe``  room supply index (base 2019=100)
    - ``Hab_Ocupadas_Caribe``     occupied rooms index (base 2019=100)
    - ``Ingreso_Real_Var_Caribe`` annual real income variation (%)

Usage::

    python scripts/extract_ema_data.py
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

RAW_DIR = Path("data/raw")
OUTPUT_DIR = Path("data/processed")
OUTPUT_FILE = OUTPUT_DIR / "master_tourism_series.csv"

MONTHS_ES: Dict[str, int] = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

# Keywords to identify target sheets across format changes.
SHEET_KEYWORDS: Dict[str, List[str]] = {
    "occupancy": ["Porc Mens Ocupac"],
    "supply_demand": ["Ind.Mes oferta.demanda"],
    "income": ["1.1", "Ing.real"],
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _clean_year(val) -> Optional[int]:
    """Parse a DANE year cell into an integer.

    Handles:

    * ``"2020(p)"``, ``"2020 (p)"``  (parenthesised provisional flag)
    * ``"2023p"``, ``"2024p"``       (bare ``p`` suffix, used in newer files)
    * ``"2019\\u1d56"``              (superscript-p)
    * plain ``int`` / ``float``
    """
    if val is None:
        return None
    raw = str(val).strip()
    raw = re.sub(
        r"[\(（]?\s*[pP]\s*[\)）]?|\u1d56", "", raw
    ).strip()
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _clean_month(val) -> Optional[str]:
    """Normalise a Spanish month name to lowercase, stripped."""
    if val is None:
        return None
    return str(val).strip().lower()


def _parse_date(year: Optional[int], month_name) -> Optional[pd.Timestamp]:
    """Build a ``pd.Timestamp`` (first of month) from year + Spanish month."""
    if year is None:
        return None
    m = _clean_month(month_name)
    if m is None:
        return None
    month_num = MONTHS_ES.get(m)
    if month_num is None:
        return None
    try:
        return pd.Timestamp(year=year, month=month_num, day=1)
    except ValueError:
        return None


def _to_float(val) -> Optional[float]:
    """Convert a cell value to float, handling ``'-'`` and empty cells."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if s in ("-", "", "N/A", "n/a"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _unaccent(text: str) -> str:
    """Return *text* with common Spanish diacritics removed."""
    return (
        text.replace("\xe1", "a")  # a-acute
        .replace("\xe9", "e")  # e-acute
        .replace("\xed", "i")  # i-acute
        .replace("\xf3", "o")  # o-acute
        .replace("\xfa", "u")  # u-acute
        .replace("\xf1", "n")  # n-tilde
    )


def _find_sheet(wb, keywords: List[str]) -> Optional[str]:
    """Return the first sheet name containing **all** *keywords* (case/accent
    insensitive)."""
    for sn in wb.sheetnames:
        sn_clean = _unaccent(sn.lower())
        if all(_unaccent(k.lower()) in sn_clean for k in keywords):
            return sn
    return None


def _find_header_row(ws) -> Optional[int]:
    """Locate the header row containing ``'Anio'`` / ``'Ano'`` in column A.

    Returns the 1-based row index, or ``None`` if not found.
    """
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None:
            s = str(val).strip().lower().replace("\xf3", "o").replace("\xf1", "n")
            if s in ("anio", "ano"):
                return row_idx
    return None


def _find_caribe_col(ws, header_row: int) -> Optional[int]:
    """Return 1-based column index whose header contains ``'caribe'``."""
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is not None and "caribe" in str(val).strip().lower():
            return col_idx
    return None


def _is_footer(a: Optional[str]) -> bool:
    """Return ``True`` if *a* looks like a footnote marker."""
    if a is None:
        return False
    s = str(a).strip()
    return bool(
        re.match(r"^[\(（]\s*[pP]", s)
        or s.lower().startswith("fuente")
        or s.lower().startswith("nota")
        or s == ""
    )


# ---------------------------------------------------------------------------
# Extractors  (each returns a DataFrame with a ``Date`` column)
# ---------------------------------------------------------------------------


def _extract_occupancy(ws) -> pd.DataFrame:
    """Extract Caribe occupancy percentage from the occupancy sheet."""
    header_row = _find_header_row(ws)
    if header_row is None:
        log.warning("Occupancy: header row not found")
        return pd.DataFrame()

    caribe_col = _find_caribe_col(ws, header_row)
    if caribe_col is None:
        log.warning("Occupancy: 'Caribe' column not found")
        return pd.DataFrame()

    records: List[Dict] = []
    current_year: Optional[int] = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        m = ws.cell(row=row_idx, column=2).value

        if _is_footer(a):
            break
        if a is not None:
            current_year = _clean_year(a)
        if current_year is None or m is None:
            continue

        date = _parse_date(current_year, m)
        if date is None:
            continue

        val = _to_float(ws.cell(row=row_idx, column=caribe_col).value)
        if val is not None:
            records.append({"Date": date, "Ocupacion_Caribe": val})

    return pd.DataFrame(records)


def _extract_supply_demand(ws) -> pd.DataFrame:
    """Extract Caribe room supply and demand indices.

    The sheet has a **two-layer header**:

    - Row 1: region names (e.g. ``Caribe``)
    - Row 2: sub-column descriptors:

      - ``Habitaciones Disponibles``
      - ``Habitaciones ocupadas``
      - ``Camas Disponibles``
      - ``Camas ocupadas``

    We extract the first two sub-columns under ``Caribe``.
    """
    header_row = _find_header_row(ws)
    if header_row is None:
        log.warning("Supply & Demand: header row not found")
        return pd.DataFrame()

    caribe_col = _find_caribe_col(ws, header_row)
    if caribe_col is None:
        log.warning("Supply & Demand: 'Caribe' column not found")
        return pd.DataFrame()

    hab_disp_col = caribe_col
    hab_ocup_col = caribe_col + 1

    records: List[Dict] = []
    current_year: Optional[int] = None

    sub_header = header_row + 1  # 2nd header level
    data_start = sub_header + 1

    for row_idx in range(data_start, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        m = ws.cell(row=row_idx, column=2).value

        if _is_footer(a):
            break
        if a is not None:
            current_year = _clean_year(a)
        if current_year is None or m is None:
            continue

        date = _parse_date(current_year, m)
        if date is None:
            continue

        hd = _to_float(ws.cell(row=row_idx, column=hab_disp_col).value)
        ho = _to_float(ws.cell(row=row_idx, column=hab_ocup_col).value)

        if hd is not None or ho is not None:
            records.append(
                {"Date": date, "Hab_Disponibles_Caribe": hd, "Hab_Ocupadas_Caribe": ho}
            )

    return pd.DataFrame(records)


def _extract_income(ws) -> pd.DataFrame:
    """Extract Caribe annual real income variation from the income sheet."""
    header_row = _find_header_row(ws)
    if header_row is None:
        log.warning("Income: header row not found")
        return pd.DataFrame()

    caribe_col = _find_caribe_col(ws, header_row)
    if caribe_col is None:
        log.warning("Income: 'Caribe' column not found")
        return pd.DataFrame()

    records: List[Dict] = []
    current_year: Optional[int] = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        m = ws.cell(row=row_idx, column=2).value

        if _is_footer(a):
            break
        if a is not None:
            current_year = _clean_year(a)
        if current_year is None or m is None:
            continue

        date = _parse_date(current_year, m)
        if date is None:
            continue

        val = _to_float(ws.cell(row=row_idx, column=caribe_col).value)
        if val is not None:
            records.append({"Date": date, "Ingreso_Real_Var_Caribe": val})

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def _extract_file(fp: Path) -> Dict[str, pd.DataFrame]:
    """Extract all three series from a single Excel file.

    Returns a dict ``{series_name: DataFrame}``.  Empty DataFrames are
    omitted.
    """
    result: Dict[str, pd.DataFrame] = {}
    wb = openpyxl.load_workbook(fp, data_only=True)

    try:
        for series_name, keywords in SHEET_KEYWORDS.items():
            sheet_name = _find_sheet(wb, keywords)
            if sheet_name is None:
                log.debug("%s: sheet not found (%s)", fp.name, keywords)
                continue

            ws = wb[sheet_name]
            if series_name == "occupancy":
                df = _extract_occupancy(ws)
            elif series_name == "supply_demand":
                df = _extract_supply_demand(ws)
            elif series_name == "income":
                df = _extract_income(ws)
            else:
                continue

            if not df.empty:
                result[series_name] = df
                log.info(
                    "%s: %s  %d rows  %s",
                    fp.name,
                    sheet_name,
                    len(df),
                    list(df.columns),
                )
    finally:
        wb.close()

    return result


def _combine_series(chunks: List[pd.DataFrame]) -> pd.DataFrame:
    """Concatenate per-file DataFrames and deduplicate by date.

    When the same date appears in multiple files the **last** occurrence
    (latest file) wins.
    """
    if not chunks:
        return pd.DataFrame()
    combined = pd.concat(chunks, ignore_index=True)
    combined = combined.drop_duplicates(subset="Date", keep="last")
    combined = combined.sort_values("Date").reset_index(drop=True)
    return combined


def main() -> None:
    """Run the full extraction pipeline."""
    files = sorted(RAW_DIR.glob("*.xlsx"))
    if not files:
        log.warning("No .xlsx files found in %s", RAW_DIR)
        return

    log.info("Found %d Excel files to process", len(files))

    occ_chunks: List[pd.DataFrame] = []
    sd_chunks: List[pd.DataFrame] = []
    inc_chunks: List[pd.DataFrame] = []

    for fp in files:
        extracted = _extract_file(fp)
        if "occupancy" in extracted:
            occ_chunks.append(extracted["occupancy"])
        if "supply_demand" in extracted:
            sd_chunks.append(extracted["supply_demand"])
        if "income" in extracted:
            inc_chunks.append(extracted["income"])

    occ = _combine_series(occ_chunks)
    sd = _combine_series(sd_chunks)
    inc = _combine_series(inc_chunks)

    log.info(
        "Series after dedup:  occupancy=%d  supply_demand=%d  income=%d",
        len(occ),
        len(sd),
        len(inc),
    )

    # Merge on Date (outer join preserves all dates)
    result = occ
    for df in (sd, inc):
        if not df.empty:
            if result.empty:
                result = df
            else:
                result = result.merge(df, on="Date", how="outer")

    if result.empty:
        log.warning("No data extracted — output file not created")
        return

    result = result.sort_values("Date").reset_index(drop=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    result.to_csv(OUTPUT_FILE, index=False)

    log.info(
        "Saved %s  (%d rows, %d columns)",
        OUTPUT_FILE,
        len(result),
        len(result.columns),
    )


if __name__ == "__main__":
    main()
